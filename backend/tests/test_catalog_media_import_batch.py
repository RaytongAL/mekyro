from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.modules.catalog import import_router
from app.scripts.seed_fake_db import IDS
from tests.conftest import auth_header


def _workbook_bytes(rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "商品名称*",
            "商品分类*",
            "商品描述",
            "SKU编码*",
            "规格_颜色*",
            "起订量*",
            "币种*",
            "库存数量*",
            "状态*",
            "阶梯价*",
            "商品图片",
            "商品详情图",
            "SKU图片",
        ]
    )
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_catalog_upload_and_product_image_lifecycle(
    client: TestClient,
    newlife_token: str,
):
    headers = auth_header(newlife_token)
    base = f"/api/v1/workspaces/{IDS['workspace_newlife']}"
    upload = client.post(
        f"{base}/uploads",
        headers=headers,
        files={"file": ("catalog-test.png", b"fake-image-content", "image/png")},
    )
    assert upload.status_code == 201
    media_url = upload.json()["url"]
    assert client.get(media_url).content == b"fake-image-content"

    gallery_ids = []
    for index in range(5):
        image = client.post(
            f"{base}/products/{IDS['product_iphone']}/images",
            headers=headers,
            json={"image_type": "product", "url": f"https://img.example.com/{index}.jpg"},
        )
        assert image.status_code == 201
        gallery_ids.append(image.json()["id"])
    too_many = client.post(
        f"{base}/products/{IDS['product_iphone']}/images",
        headers=headers,
        json={"image_type": "product", "url": "https://img.example.com/overflow.jpg"},
    )
    assert too_many.status_code == 409

    first_detail = client.post(
        f"{base}/products/{IDS['product_iphone']}/images",
        headers=headers,
        json={"image_type": "product_detail", "url": media_url},
    )
    second_detail = client.post(
        f"{base}/products/{IDS['product_iphone']}/images",
        headers=headers,
        json={"image_type": "product_detail", "url": "https://img.example.com/detail.jpg"},
    )
    assert first_detail.status_code == 201
    assert second_detail.status_code == 201

    sku_image = client.post(
        f"{base}/products/{IDS['product_iphone']}/images",
        headers=headers,
        json={
            "image_type": "sku",
            "variant_id": IDS["variant_iphone_a"],
            "url": "https://img.example.com/sku.jpg",
        },
    )
    assert sku_image.status_code == 201
    product = client.get(f"{base}/products/{IDS['product_iphone']}", headers=headers).json()
    detail_images = [item for item in product["images"] if item["image_type"] == "product_detail"]
    assert [item["id"] for item in detail_images] == [second_detail.json()["id"]]
    variant = next(item for item in product["variants"] if item["id"] == IDS["variant_iphone_a"])
    assert [item["id"] for item in variant["images"]] == [sku_image.json()["id"]]

    deleted = client.delete(
        f"{base}/products/{IDS['product_iphone']}/images/{gallery_ids[0]}",
        headers=headers,
    )
    assert deleted.status_code == 204


def test_product_image_rejects_cross_product_variant(
    client: TestClient,
    newlife_token: str,
):
    response = client.post(
        f"/api/v1/workspaces/{IDS['workspace_newlife']}/products/{IDS['product_iphone']}/images",
        headers=auth_header(newlife_token),
        json={
            "image_type": "sku",
            "variant_id": IDS["variant_charger"],
            "url": "https://img.example.com/wrong-product.jpg",
        },
    )
    assert response.status_code == 404


def test_product_import_template_preview_and_confirm(
    client: TestClient,
    newlife_token: str,
):
    headers = auth_header(newlife_token)
    base = f"/api/v1/workspaces/{IDS['workspace_newlife']}"
    template = client.get(f"{base}/product-import/template", headers=headers)
    assert template.status_code == 200
    assert template.content.startswith(b"PK")
    template_workbook = load_workbook(BytesIO(template.content))
    template_sheet = template_workbook["商品导入模板"]
    assert template_workbook.sheetnames == ["商品导入模板"]
    assert template_sheet["A3"].value == "商品名称*"
    assert template_sheet["O3"].value == "SKU图片"
    assert template_sheet.max_row == 4
    template_workbook.close()

    template_rows, template_errors, template_total = import_router._parse_workbook(template.content)
    assert template_total == 1
    assert len(template_rows) == 1
    assert template_errors == []

    workbook = _workbook_bytes(
        [
            [
                "Imported Travel Adapter",
                "Imported/Travel",
                "International adapter",
                "IMPORT-ADAPTER-EU",
                "Black",
                10,
                "USD",
                40,
                "active",
                12.5,
                "https://img.example.com/import-product.jpg",
                "https://img.example.com/import-detail.jpg",
                "https://img.example.com/import-eu.jpg",
            ],
            [
                "Imported Travel Adapter",
                "Imported/Travel",
                "International adapter",
                "IMPORT-ADAPTER-UK",
                "White",
                20,
                "USD",
                25,
                "active",
                11.0,
                "",
                "",
                "https://img.example.com/import-uk.jpg",
            ],
        ]
    )
    preview = client.post(
        f"{base}/product-import/preview",
        headers=headers,
        files={
            "file": (
                "catalog.xlsx",
                workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200
    assert preview.json()["summary"] == {
        "total": 2,
        "valid": 2,
        "invalid": 0,
        "products": 1,
        "skus": 2,
    }

    confirm = client.post(
        f"{base}/product-import/confirm",
        headers=headers,
        json={"rows": preview.json()["rows"]},
    )
    assert confirm.status_code == 201
    assert confirm.json() == {
        "created_products": 1,
        "created_skus": 2,
        "batches": 1,
        "errors": [],
    }
    products = client.get(f"{base}/products?search=Imported%20Travel", headers=headers).json()
    assert products["total"] == 1
    assert len(products["items"][0]["variants"]) == 2
    assert len(products["items"][0]["images"]) == 2

    duplicate = client.post(
        f"{base}/product-import/confirm",
        headers=headers,
        json={"rows": preview.json()["rows"]},
    )
    assert duplicate.status_code == 409


def test_product_import_preview_reports_invalid_rows(
    client: TestClient,
    newlife_token: str,
):
    workbook = _workbook_bytes(
        [["", "Imported", "", "", "", 0, "XXX", -1, "broken", 0, "", "", ""]]
    )
    response = client.post(
        f"/api/v1/workspaces/{IDS['workspace_newlife']}/product-import/preview",
        headers=auth_header(newlife_token),
        files={"file": ("invalid.xlsx", workbook, "application/octet-stream")},
    )
    assert response.status_code == 200
    assert response.json()["summary"]["valid"] == 0
    assert response.json()["summary"]["invalid"] == 1
    assert response.json()["errors"][0]["errors"]


def test_product_import_accepts_valid_sku_without_specification_template(
    client: TestClient,
    newlife_token: str,
):
    workbook = _workbook_bytes(
        [["No Spec Product", "Imports", "", "NO-SPEC-001", "", 1, "USD", 5, "active", 9.9, "", "", ""]]
    )
    response = client.post(
        f"/api/v1/workspaces/{IDS['workspace_newlife']}/product-import/preview",
        headers=auth_header(newlife_token),
        files={"file": ("no-spec.xlsx", workbook, "application/octet-stream")},
    )
    assert response.status_code == 200
    assert response.json()["summary"]["valid"] == 1
    assert response.json()["summary"]["invalid"] == 0
    assert response.json()["rows"][0]["specs"] == {}


def _import_row(row: int, sku_code: str, product_name: str = "Batched Product") -> dict:
    return {
        "row": row,
        "product_name": product_name,
        "category_path": "Batch Imports/Test",
        "description": "Batch transaction test",
        "sku_code": sku_code,
        "specs": {"Color": f"Color {row}"},
        "moq": 1,
        "currency": "USD",
        "stock_quantity": row,
        "status": "active",
        "unit_price": 10 + row,
        "product_images": [],
        "product_detail_image": "",
        "sku_image": "",
    }


def test_product_import_uses_django_compatible_independent_batches(
    client: TestClient,
    newlife_token: str,
    monkeypatch,
):
    monkeypatch.setattr(import_router, "IMPORT_BATCH_SIZE", 2)
    base = f"/api/v1/workspaces/{IDS['workspace_newlife']}"
    response = client.post(
        f"{base}/product-import/confirm",
        headers=auth_header(newlife_token),
        json={
            "rows": [
                _import_row(1, "BATCH-SPLIT-001"),
                _import_row(2, "BATCH-SPLIT-002"),
                _import_row(3, "BATCH-SPLIT-003"),
            ]
        },
    )
    assert response.status_code == 201, response.text
    assert response.json() == {
        "created_products": 2,
        "created_skus": 3,
        "batches": 2,
        "errors": [],
    }
    products = client.get(
        f"{base}/products?search=Batched%20Product",
        headers=auth_header(newlife_token),
    ).json()
    assert products["total"] == 2
    assert sorted(len(item["variants"]) for item in products["items"]) == [1, 2]


def test_product_import_preserves_committed_batches_when_a_later_batch_fails(
    client: TestClient,
    newlife_token: str,
    monkeypatch,
):
    monkeypatch.setattr(import_router, "IMPORT_BATCH_SIZE", 2)
    original = import_router._import_product_batch
    calls = 0

    async def fail_second_batch(rows, context, session):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("private database detail")
        return await original(rows, context, session)

    monkeypatch.setattr(import_router, "_import_product_batch", fail_second_batch)
    base = f"/api/v1/workspaces/{IDS['workspace_newlife']}"
    response = client.post(
        f"{base}/product-import/confirm",
        headers=auth_header(newlife_token),
        json={
            "rows": [
                _import_row(1, "BATCH-PARTIAL-001", "Committed Batch A"),
                _import_row(2, "BATCH-PARTIAL-002", "Committed Batch B"),
                _import_row(3, "BATCH-PARTIAL-003", "Failed Batch"),
            ]
        },
    )
    assert response.status_code == 201, response.text
    assert response.json() == {
        "created_products": 2,
        "created_skus": 2,
        "batches": 2,
        "errors": ["Batch 2 failed due to write failure"],
    }
    assert "private database detail" not in response.text
    products = client.get(
        f"{base}/products?search=Batch",
        headers=auth_header(newlife_token),
    ).json()
    names = {item["name"] for item in products["items"]}
    assert names >= {"Committed Batch A", "Committed Batch B"}
    assert "Failed Batch" not in names


def test_catalog_batch_commands_are_atomic_and_idempotent(
    client: TestClient,
    newlife_token: str,
):
    headers = auth_header(newlife_token)
    base = f"/api/v1/workspaces/{IDS['workspace_newlife']}"
    variants = [IDS["variant_iphone_a"], IDS["variant_iphone_b"]]
    category = client.post(
        f"{base}/categories",
        headers=headers,
        json={"name": "Batch SKU Parent Category"},
    ).json()
    updated = client.patch(
        f"{base}/batch/variants",
        headers=headers,
        json=[
            {
                "id": variants[0],
                "minimum_order_quantity": 15,
                "stock_quantity": 90,
                "product_name": "Batch Updated iPhone",
                "product_category_id": category["id"],
            },
            {"id": variants[1], "status": "inactive"},
        ],
    )
    assert updated.status_code == 200
    assert updated.json()[0]["minimum_order_quantity"] == 15
    assert updated.json()[0]["stock_quantity"] == 90
    assert updated.json()[1]["status"] == "inactive"
    product = client.get(
        f"{base}/products/{IDS['product_iphone']}", headers=headers
    ).json()
    assert product["name"] == "Batch Updated iPhone"
    assert product["category_id"] == category["id"]

    foreign_batch = client.patch(
        f"{base}/batch/variants",
        headers=headers,
        json=[
            {"id": variants[0], "minimum_order_quantity": 99},
            {"id": IDS["variant_lamp"], "status": "inactive"},
        ],
    )
    assert foreign_batch.status_code == 404
    unchanged = client.get(f"{base}/variants/{variants[0]}", headers=headers)
    assert unchanged.json()["minimum_order_quantity"] == 15

    foreign_category = client.patch(
        f"{base}/batch/variants",
        headers=headers,
        json=[
            {
                "id": variants[0],
                "stock_quantity": 1,
                "product_category_id": IDS["category_home"],
            }
        ],
    )
    assert foreign_category.status_code == 404
    unchanged_stock = client.get(f"{base}/variants/{variants[0]}", headers=headers)
    assert unchanged_stock.json()["stock_quantity"] == 90

    prices = client.put(
        f"{base}/batch/price-tiers",
        headers=headers,
        json=[
            {
                "variant_id": variants[0],
                "price_tiers": [{"minimum_quantity": 15, "unit_price": "399.00"}],
            },
            {
                "variant_id": variants[1],
                "price_tiers": [{"minimum_quantity": 20, "unit_price": "349.00"}],
            },
        ],
    )
    assert prices.status_code == 200
    assert [item["variant_id"] for item in prices.json()] == variants

    batch_headers = {**headers, "Idempotency-Key": "catalog-batch-inventory-001"}
    inventory_request = [
        {
            "variant_id": variants[0],
            "quantity_delta": 2,
            "reason": "Batch cycle count",
            "reference": "BATCH-COUNT-1",
        },
        {
            "variant_id": variants[0],
            "quantity_delta": -1,
            "reason": "Batch correction",
            "reference": "BATCH-COUNT-2",
        },
        {
            "variant_id": variants[1],
            "quantity_delta": 3,
            "reason": "Batch receipt",
            "reference": "BATCH-COUNT-3",
        },
    ]
    first = client.post(
        f"{base}/batch/inventory-adjustments",
        headers=batch_headers,
        json=inventory_request,
    )
    replay = client.post(
        f"{base}/batch/inventory-adjustments",
        headers=batch_headers,
        json=inventory_request,
    )
    assert first.status_code == 201
    assert replay.status_code == 201
    assert replay.json() == first.json()
    assert [item["balance_after"] for item in first.json()["items"]] == [92, 91, 39]

    conflict = client.post(
        f"{base}/batch/inventory-adjustments",
        headers=batch_headers,
        json=[{**inventory_request[0], "quantity_delta": 5}],
    )
    assert conflict.status_code == 409
