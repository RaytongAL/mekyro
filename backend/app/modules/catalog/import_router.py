from collections import defaultdict
from io import BytesIO
from typing import Annotated, Literal

import anyio
from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field, ValidationError, model_validator
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError

from app.core.audit import record_audit
from app.core.dependencies import SessionDep, WorkspaceDep, WorkspaceWriteDep
from app.core.models import Category, PriceTier, Product, ProductImage, ProductVariant, new_id
from app.modules.catalog.reference_template import reference_template_bytes
from app.modules.catalog.sync_outbox import enqueue_catalog_sync_if_configured

router = APIRouter(prefix="/workspaces/{workspace_id}/product-import", tags=["catalog-import"])

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
IMPORT_BATCH_SIZE = 200
FIXED_COLUMNS = {
    "商品名称": "product_name",
    "商品分类": "category_path",
    "商品描述": "description",
    "SKU编码": "sku_code",
    "起订量": "moq",
    "币种": "currency",
    "库存数量": "stock_quantity",
    "状态": "status",
    "阶梯价": "unit_price",
    "商品图片": "product_images",
    "商品详情图": "product_detail_image",
    "SKU图片": "sku_image",
}


class ProductImportRow(BaseModel):
    row: int = Field(ge=1)
    product_name: str = Field(min_length=1, max_length=200)
    category_path: str = Field(min_length=1, max_length=1000)
    description: str = Field(default="", max_length=10000)
    sku_code: str = Field(min_length=1, max_length=100)
    specs: dict[str, str]
    moq: int = Field(default=1, ge=1, le=1_000_000_000)
    currency: Literal["USD", "CNY", "EUR", "GBP"] = "USD"
    stock_quantity: int = Field(default=0, ge=0, le=1_000_000_000)
    status: Literal["active", "inactive"] = "active"
    unit_price: float = Field(gt=0)
    product_images: list[str] = Field(default_factory=list, max_length=5)
    product_detail_image: str = Field(default="", max_length=2000)
    sku_image: str = Field(default="", max_length=2000)

    @model_validator(mode="after")
    def validate_business_fields(self):
        parts = [part.strip() for part in self.category_path.split("/") if part.strip()]
        if not parts or len(parts) > 5:
            raise ValueError("Category path must contain between one and five levels")
        return self


class ImportError(BaseModel):
    row: int
    sku_code: str = ""
    errors: list[str]


class ImportSummary(BaseModel):
    total: int
    valid: int
    invalid: int
    products: int
    skus: int


class ImportPreviewResponse(BaseModel):
    rows: list[ProductImportRow]
    errors: list[ImportError]
    summary: ImportSummary


class ImportConfirmRequest(BaseModel):
    rows: list[ProductImportRow] = Field(min_length=1, max_length=MAX_IMPORT_ROWS)


class ImportConfirmResponse(BaseModel):
    created_products: int
    created_skus: int
    batches: int
    errors: list[str]


def _template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品导入"
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="163F3B")
    required_fill = PatternFill("solid", fgColor="276D61")
    optional_fill = PatternFill("solid", fgColor="64748B")
    note_fill = PatternFill("solid", fgColor="E8F2EC")
    sample_fill = PatternFill("solid", fgColor="F7FAF8")
    white_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
    body_font = Font(name="Microsoft YaHei", color="24312F", size=10)
    thin_gray = Side(style="thin", color="D8E1DE")

    headers = [
        "商品名称*",
        "商品分类*",
        "商品描述",
        "SKU编码*",
        "规格_颜色",
        "规格_型号",
        "起订量",
        "币种",
        "库存数量",
        "状态",
        "阶梯价*",
        "商品图片",
        "商品详情图",
        "SKU图片",
    ]
    header_notes = [
        "必填。同一商品有多个 SKU 时，每行重复填写相同商品名称。最长 200 个字符。",
        "必填。使用 / 分隔分类层级，最多 5 级，例如：电子产品/充电设备。不存在的分类会自动创建。",
        "选填。商品介绍，最长 10000 个字符。",
        "必填。每个 SKU 编码必须唯一，最长 100 个字符。建议按文本填写，避免前导零丢失。",
        "选填规格。可把“颜色”改成实际规格名，也可新增以“规格_”开头的列。",
        "选填规格。可把“型号”改成容量、尺寸、材质等实际规格名。",
        "选填，默认为 1。必须是大于等于 1 的整数。",
        "选填，默认为 USD。可选 USD、CNY、EUR、GBP。",
        "选填，默认为 0。必须是大于等于 0 的整数。",
        "选填，默认为 active。active 表示启用，inactive 表示停用。",
        "必填。必须是大于 0 的数字，当前导入支持每个 SKU 一个基础价格。",
        "选填。填写公开可访问的图片 URL；多张用英文逗号分隔，最多 5 张。",
        "选填。填写 1 个公开可访问的商品详情图 URL。",
        "选填。填写 1 个公开可访问的 SKU 图片 URL。",
    ]

    sheet.merge_cells("A1:N1")
    sheet["A1"] = "Mekyro 商品批量导入模板"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:N2")
    sheet["A2"] = "从第 4 行开始填写；带 * 为必填；每个 SKU 占一行；同一商品的多个 SKU 请重复商品信息。详细规则见“填写说明”。"
    sheet["A2"].fill = note_fill
    sheet["A2"].font = Font(name="Microsoft YaHei", color="31584F", italic=True, size=10)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 32

    for column, (header, note) in enumerate(zip(headers, header_notes, strict=True), start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.fill = required_fill if header.endswith("*") else optional_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
        cell.comment = Comment(note, "Mekyro")
    sheet.row_dimensions[3].height = 30

    sample_rows = [
        [
            "便携式氮化镓充电器",
            "电子产品/充电设备",
            "适用于跨境批发的 65W 多口快充",
            "CHARGER-65W-BLACK",
            "黑色",
            "65W",
            10,
            "USD",
            100,
            "active",
            18.9,
            "https://example.com/charger-front.jpg,https://example.com/charger-side.jpg",
            "https://example.com/charger-detail.jpg",
            "https://example.com/charger-black.jpg",
        ],
        [
            "便携式氮化镓充电器",
            "电子产品/充电设备",
            "适用于跨境批发的 65W 多口快充",
            "CHARGER-65W-WHITE",
            "白色",
            "65W",
            20,
            "USD",
            60,
            "active",
            18.5,
            "",
            "",
            "https://example.com/charger-white.jpg",
        ],
    ]
    for row_index, values in enumerate(sample_rows, start=4):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = body_font
            cell.fill = sample_fill if row_index % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            cell.alignment = Alignment(
                horizontal="center" if column in {5, 6, 7, 8, 9, 10, 11} else "left",
                vertical="top",
                wrap_text=column in {1, 2, 3, 12, 13, 14},
            )
            cell.border = Border(bottom=thin_gray)
        sheet.row_dimensions[row_index].height = 38

    widths = [24, 24, 34, 24, 14, 14, 12, 10, 12, 12, 12, 42, 38, 38]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=3, column=column).column_letter].width = width

    sheet.freeze_panes = "D4"
    sheet.auto_filter.ref = f"A3:N{MAX_IMPORT_ROWS + 3}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:3"
    sheet["D4"].number_format = "@"
    sheet["D5"].number_format = "@"
    sheet["K4"].number_format = "0.00"
    sheet["K5"].number_format = "0.00"

    validations = [
        (DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True), f"G4:G{MAX_IMPORT_ROWS + 3}", "起订量必须是大于等于 1 的整数"),
        (DataValidation(type="list", formula1='"USD,CNY,EUR,GBP"', allow_blank=True), f"H4:H{MAX_IMPORT_ROWS + 3}", "请选择支持的币种"),
        (DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True), f"I4:I{MAX_IMPORT_ROWS + 3}", "库存数量必须是大于等于 0 的整数"),
        (DataValidation(type="list", formula1='"active,inactive"', allow_blank=True), f"J4:J{MAX_IMPORT_ROWS + 3}", "请选择 active 或 inactive"),
        (DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False), f"K4:K{MAX_IMPORT_ROWS + 3}", "阶梯价必须是大于 0 的数字"),
    ]
    for validation, cell_range, error_message in validations:
        validation.error = error_message
        validation.errorTitle = "填写格式不正确"
        validation.promptTitle = "填写提示"
        validation.prompt = error_message
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(cell_range)

    guide = workbook.create_sheet("填写说明")
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:E1")
    guide["A1"] = "商品导入填写说明"
    guide["A1"].fill = title_fill
    guide["A1"].font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=15)
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 32
    guide.merge_cells("A2:E2")
    guide["A2"] = "建议先保留示例行理解格式，再替换为实际商品数据。单次最多导入 5000 行，文件大小不超过 10 MB。"
    guide["A2"].fill = note_fill
    guide["A2"].font = body_font
    guide["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    guide.row_dimensions[2].height = 30
    guide_headers = ["字段", "是否必填", "格式 / 可选值", "示例", "说明"]
    guide.append([])
    guide.append(guide_headers)
    guide_rows = [
        ["商品名称", "是", "文本，最长 200 字符", "便携式氮化镓充电器", "同一商品的多个 SKU 必须重复填写相同商品名称。"],
        ["商品分类", "是", "1-5 级，用 / 分隔", "电子产品/充电设备", "分类不存在时系统会自动创建。"],
        ["商品描述", "否", "文本，最长 10000 字符", "65W 多口快充", "同一商品建议保持一致。"],
        ["SKU编码", "是", "唯一文本，最长 100 字符", "CHARGER-65W-BLACK", "不要使用已存在的 SKU 编码；前导零会按文本保留。"],
        ["规格_字段名", "否", "列名必须以 规格_ 开头", "规格_颜色 = 黑色", "可新增多个规格列，例如规格_尺寸、规格_材质。"],
        ["起订量", "否", "整数，>= 1；默认 1", 10, "表示该 SKU 的最小订购数量。"],
        ["币种", "否", "USD / CNY / EUR / GBP；默认 USD", "USD", "使用大写币种代码。"],
        ["库存数量", "否", "整数，>= 0；默认 0", 100, "导入后的初始库存。"],
        ["状态", "否", "active / inactive；默认 active", "active", "active 为启用，inactive 为停用。"],
        ["阶梯价", "是", "数字，> 0", 18.9, "当前导入支持每个 SKU 一个基础价格。"],
        ["商品图片", "否", "公开 URL，最多 5 张", "https://example.com/a.jpg", "多张图片使用英文逗号分隔。"],
        ["商品详情图", "否", "1 个公开 URL", "https://example.com/detail.jpg", "同一商品建议只在第一条 SKU 行填写。"],
        ["SKU图片", "否", "1 个公开 URL", "https://example.com/sku.jpg", "对应当前行的 SKU。"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[4]:
        cell.fill = required_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in guide.iter_rows(min_row=5, max_row=4 + len(guide_rows), min_col=1, max_col=5):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
    for column, width in enumerate([20, 12, 34, 34, 48], start=1):
        guide.column_dimensions[guide.cell(row=4, column=column).column_letter].width = width
    for row_index in range(5, 5 + len(guide_rows)):
        guide.row_dimensions[row_index].height = 34
    guide.freeze_panes = "A5"
    guide.auto_filter.ref = f"A4:E{4 + len(guide_rows)}"

    workbook.active = 0
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _parse_workbook(data: bytes) -> tuple[list[ProductImportRow], list[ImportError], int]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    header_index: int | None = None
    header_map: dict[int, str] = {}
    spec_columns: dict[int, str] = {}
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=10, values_only=True),
        start=1,
    ):
        values = [str(value).strip() if value is not None else "" for value in row]
        if not any("商品名称" in value or "SKU编码" in value for value in values):
            continue
        header_index = row_index
        for column_index, raw_value in enumerate(values):
            value = raw_value.rstrip("*")
            if value in FIXED_COLUMNS:
                header_map[column_index] = FIXED_COLUMNS[value]
            elif value.startswith("规格_") and value[3:]:
                spec_columns[column_index] = value[3:]
        break
    if header_index is None:
        workbook.close()
        return [], [ImportError(row=0, errors=["Template header row was not found"])], 0

    raw_rows = [
        row
        for row in sheet.iter_rows(min_row=header_index + 1, values_only=True)
        if row and any(value is not None and str(value).strip() for value in row)
    ]
    if len(raw_rows) > MAX_IMPORT_ROWS:
        workbook.close()
        return (
            [],
            [ImportError(row=0, errors=[f"Import cannot exceed {MAX_IMPORT_ROWS} rows"])],
            len(raw_rows),
        )

    valid_rows: list[ProductImportRow] = []
    errors: list[ImportError] = []
    seen_sku_codes: set[str] = set()
    for logical_row, row in enumerate(raw_rows, start=1):
        parsed: dict = {
            "row": logical_row,
            "product_name": "",
            "category_path": "",
            "description": "",
            "sku_code": "",
            "specs": {},
            "moq": 1,
            "currency": "USD",
            "stock_quantity": 0,
            "status": "active",
            "unit_price": 0,
            "product_images": [],
            "product_detail_image": "",
            "sku_image": "",
        }
        for column_index, field_name in header_map.items():
            value = row[column_index] if column_index < len(row) else None
            string_value = str(value).strip() if value is not None else ""
            if field_name in {"moq", "stock_quantity"}:
                parsed[field_name] = value if value not in {None, ""} else parsed[field_name]
            elif field_name == "unit_price":
                parsed[field_name] = value if value not in {None, ""} else 0
            elif field_name == "currency":
                parsed[field_name] = string_value.upper() or "USD"
            elif field_name == "status":
                parsed[field_name] = string_value.lower() or "active"
            elif field_name == "product_images":
                parsed[field_name] = [
                    item.strip() for item in string_value.split(",") if item.strip()
                ][:5]
            else:
                parsed[field_name] = string_value
        for column_index, spec_name in spec_columns.items():
            value = row[column_index] if column_index < len(row) else None
            string_value = str(value).strip() if value is not None else ""
            if string_value:
                parsed["specs"][spec_name] = string_value
        try:
            item = ProductImportRow.model_validate(parsed)
            if item.sku_code in seen_sku_codes:
                raise ValueError("SKU code is duplicated in the workbook")
            seen_sku_codes.add(item.sku_code)
            valid_rows.append(item)
        except (ValidationError, ValueError) as exc:
            messages = (
                [str(error["msg"]) for error in exc.errors()]
                if isinstance(exc, ValidationError)
                else [str(exc)]
            )
            errors.append(
                ImportError(
                    row=logical_row,
                    sku_code=str(parsed.get("sku_code", "")),
                    errors=messages,
                )
            )
    workbook.close()
    return valid_rows, errors, len(raw_rows)


async def _resolve_category_path(
    workspace_id: str,
    category_path: str,
    session: SessionDep,
) -> Category:
    parent_id: str | None = None
    category: Category | None = None
    for name in [part.strip() for part in category_path.split("/") if part.strip()]:
        parent_filter = (
            Category.parent_id == parent_id
            if parent_id is not None
            else Category.parent_id.is_(None)
        )
        category = await session.scalar(
            select(Category).where(
                Category.workspace_id == workspace_id,
                Category.name == name,
                parent_filter,
            )
        )
        if category is None:
            category = Category(
                id=new_id(),
                workspace_id=workspace_id,
                parent_id=parent_id,
                name=name,
            )
            session.add(category)
            await session.flush()
        parent_id = category.id
    if category is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Category path is empty",
        )
    return category


@router.get("/template")
async def download_product_import_template(context: WorkspaceDep) -> StreamingResponse:
    del context
    content = await anyio.to_thread.run_sync(reference_template_bytes)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="product_import_template.xlsx"'},
    )


@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_product_import(
    context: WorkspaceWriteDep,
    session: SessionDep,
    file: Annotated[UploadFile, File()],
) -> ImportPreviewResponse:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Only .xlsx files are supported",
        )
    data = await file.read(MAX_IMPORT_BYTES + 1)
    await file.close()
    if len(data) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Import file exceeds 10 MB",
        )
    try:
        rows, errors, total = await anyio.to_thread.run_sync(_parse_workbook, data)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="The workbook could not be parsed",
        ) from exc

    if rows:
        existing_codes = set(
            await session.scalars(
                select(ProductVariant.sku_code).where(
                    ProductVariant.workspace_id == context.workspace.id,
                    ProductVariant.sku_code.in_([item.sku_code for item in rows]),
                )
            )
        )
        accepted: list[ProductImportRow] = []
        for item in rows:
            if item.sku_code in existing_codes:
                errors.append(
                    ImportError(
                        row=item.row,
                        sku_code=item.sku_code,
                        errors=["SKU code already exists in this Workspace"],
                    )
                )
            else:
                accepted.append(item)
        rows = accepted
    return ImportPreviewResponse(
        rows=rows,
        errors=errors,
        summary=ImportSummary(
            total=total,
            valid=len(rows),
            invalid=len(errors),
            products=len({item.product_name for item in rows}),
            skus=len(rows),
        ),
    )


async def _import_product_batch(
    batch_rows: list[ProductImportRow],
    context: WorkspaceWriteDep,
    session: SessionDep,
) -> tuple[int, int, list[str]]:
    groups: dict[str, list[ProductImportRow]] = defaultdict(list)
    created_product_ids: list[str] = []
    for row in batch_rows:
        groups[row.product_name].append(row)
    for product_name, product_rows in groups.items():
        first = product_rows[0]
        category = await _resolve_category_path(
            context.workspace.id,
            first.category_path,
            session,
        )
        specification_values: dict[str, set[str]] = defaultdict(set)
        for row in product_rows:
            for key, value in row.specs.items():
                if value:
                    specification_values[key].add(value)
        product = Product(
            id=new_id(),
            workspace_id=context.workspace.id,
            category_id=category.id,
            name=product_name,
            description=first.description,
            specification_template=[
                {"name": key, "options": sorted(values)}
                for key, values in specification_values.items()
            ],
            status=first.status,
        )
        created_product_ids.append(product.id)
        session.add(product)
        for url in first.product_images:
            session.add(
                ProductImage(
                    id=new_id(),
                    product_id=product.id,
                    image_type="product",
                    file_key=url,
                )
            )
        if first.product_detail_image:
            session.add(
                ProductImage(
                    id=new_id(),
                    product_id=product.id,
                    image_type="product_detail",
                    file_key=first.product_detail_image,
                )
            )
        for row in product_rows:
            variant = ProductVariant(
                id=new_id(),
                workspace_id=context.workspace.id,
                product_id=product.id,
                sku_code=row.sku_code,
                specifications=row.specs,
                minimum_order_quantity=row.moq,
                currency=row.currency,
                stock_quantity=row.stock_quantity,
                status=row.status,
            )
            session.add(variant)
            session.add(
                PriceTier(
                    id=new_id(),
                    variant_id=variant.id,
                    minimum_quantity=row.moq,
                    unit_price=row.unit_price,
                )
            )
            if row.sku_image:
                session.add(
                    ProductImage(
                        id=new_id(),
                        variant_id=variant.id,
                        image_type="sku",
                        file_key=row.sku_image,
                    )
                )
    record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        action="catalog.products_imported",
        entity_type="workspace",
        entity_id=context.workspace.id,
        payload={"product_count": len(groups), "variant_count": len(batch_rows)},
    )
    await enqueue_catalog_sync_if_configured(
        session,
        workspace_id=context.workspace.id,
        operation="catalog",
        product_ids=created_product_ids,
    )
    return len(groups), len(batch_rows), created_product_ids


@router.post("/confirm", response_model=ImportConfirmResponse, status_code=status.HTTP_201_CREATED)
async def confirm_product_import(
    payload: ImportConfirmRequest,
    context: WorkspaceWriteDep,
    session: SessionDep,
) -> ImportConfirmResponse:
    sku_codes = [item.sku_code for item in payload.rows]
    if len(sku_codes) != len(set(sku_codes)):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Import rows contain duplicate SKU codes",
        )
    existing = set(
        await session.scalars(
            select(ProductVariant.sku_code).where(
                ProductVariant.workspace_id == context.workspace.id,
                ProductVariant.sku_code.in_(sku_codes),
            )
        )
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"message": "SKU codes already exist", "sku_codes": sorted(existing)},
        )

    created_products = 0
    created_skus = 0
    batches = 0
    errors: list[str] = []
    for offset in range(0, len(payload.rows), IMPORT_BATCH_SIZE):
        batches += 1
        batch = payload.rows[offset : offset + IMPORT_BATCH_SIZE]
        try:
            product_count, sku_count, _product_ids = await _import_product_batch(
                batch,
                context,
                session,
            )
            await session.commit()
            created_products += product_count
            created_skus += sku_count
        except Exception as exc:
            await session.rollback()
            error_kind = "catalog conflict" if isinstance(exc, IntegrityError) else "write failure"
            errors.append(f"Batch {batches} failed due to {error_kind}")
    return ImportConfirmResponse(
        created_products=created_products,
        created_skus=created_skus,
        batches=batches,
        errors=errors,
    )
