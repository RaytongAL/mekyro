from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def reference_template_bytes() -> bytes:
    """Generate the template used by the reference import service."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品导入模板"

    title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet.merge_cells("A1:O1")
    sheet["A1"] = "📦 商品批量导入模板"
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 36

    sheet.merge_cells("A2:O2")
    sheet["A2"] = (
        "填写说明："
        "① 带 * 的列为必填；"
        "② 商品分类用 / 分隔层级（如 手机配件/保护壳），不存在则自动创建；"
        "③ 规格至少填一个（如规格_颜色），可自由增减列；"
        "④ 阶梯价为单价数字，如 5.50 表示 ≥ 起订量时单价 $5.50；"
        "⑤ 币种可选 USD/CNY/EUR/GBP；"
        "⑥ 商品图片多张用逗号分隔，最大 5 张；"
        "⑦ 状态可选 active（启用）/ inactive（停用）；"
        "⑧ 同名商品自动合并，SKU 追加到该商品中"
    )
    sheet["A2"].font = Font(name="Microsoft YaHei", size=9, color="888888")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 44

    headers = [
        "商品名称*", "商品分类*", "商品描述", "SKU编码*",
        "规格_颜色*", "规格_容量", "规格_版本", "起订量*", "币种*",
        "库存数量*", "状态*", "阶梯价*", "商品图片", "商品详情图", "SKU图片",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    sheet.row_dimensions[3].height = 32

    examples = [
        ["iPhone 15 手机壳", "手机配件/保护壳", "液态硅胶防摔", "IP15-CASE-BLK", "黑色", "", "国际版", 1, "USD", 500, "active", 5.50, "https://img.example.com/case-1.jpg, https://img.example.com/case-2.jpg", "https://img.example.com/case-detail.jpg", "https://img.example.com/case-blk.jpg"],
    ]
    example_font = Font(name="Microsoft YaHei", size=10, color="666666")
    for row_index, example in enumerate(examples, start=4):
        for column, value in enumerate(example, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = example_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    widths = [22, 18, 25, 18, 12, 12, 12, 10, 8, 10, 10, 14, 36, 30, 28]
    for column, width in enumerate(widths, start=1):
        column_letter = sheet.cell(row=3, column=column).column_letter
        sheet.column_dimensions[column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
