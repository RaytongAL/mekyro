from collections import defaultdict
from io import BytesIO
from typing import Annotated, Literal

import anyio
from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, ValidationError, model_validator
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError

from app.core.audit import record_audit
from app.core.dependencies import SessionDep, WorkspaceDep, WorkspaceWriteDep
from app.core.models import Category, PriceTier, Product, ProductImage, ProductVariant, new_id
from app.modules.shopify.outbox import enqueue_shopify_sync_if_active

router = APIRouter(prefix="/workspaces/{workspace_id}/product-import", tags=["catalog-import"])

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
IMPORT_BATCH_SIZE = 200
FIXED_COLUMNS = {
    "商品名称": "product_name",
    "商品分类": "category_path",
    "商品描述": "description",
    "SKU编码": "sku_code",
    "起订量": "moq",
    "币种": "currency",
    "库存数量": "stock_quantity",
    "状态": "status",
    "阶梯价": "unit_price",
    "商品图片": "product_images",
    "商品详情图": "product_detail_image",
    "SKU图片": "sku_image",
}


class ProductImportRow(BaseModel):
    row: int = Field(ge=1)
    product_name: str = Field(min_length=1, max_length=200)
    category_path: str = Field(min_length=1, max_length=1000)
    description: str = Field(default="", max_length=10000)
    sku_code: str = Field(min_length=1, max_length=100)
    specs: dict[str, str]
    moq: int = Field(default=1, ge=1, le=1_000_000_000)
    currency: Literal["USD", "CNY", "EUR", "GBP"] = "USD"
    stock_quantity: int = Field(default=0, ge=0, le=1_000_000_000)
    status: Literal["active", "inactive"] = "active"
    unit_price: float = Field(gt=0)
    product_images: list[str] = Field(default_factory=list, max_length=5)
    product_detail_image: str = Field(default="", max_length=2000)
    sku_image: str = Field(default="", max_length=2000)

    @model_validator(mode="after")
    def validate_business_fields(self):
        parts = [part.strip() for part in self.category_path.split("/") if part.strip()]
        if not parts or len(parts) > 5:
            raise ValueError("Category path must contain between one and five levels")
        return self


class ImportError(BaseModel):
    row: int
    sku_code: str = ""
    errors: list[str]


class ImportSummary(BaseModel):
    total: int
    valid: int
    invalid: int
    products: int
    skus: int


class ImportPreviewResponse(BaseModel):
    rows: list[ProductImportRow]
    errors: list[ImportError]
    summary: ImportSummary


class ImportConfirmRequest(BaseModel):
    rows: list[ProductImportRow] = Field(min_length=1, max_length=MAX_IMPORT_ROWS)


class ImportConfirmResponse(BaseModel):
    created_products: int
    created_skus: int
    batches: int
    errors: list[str]


def _template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品导入模板"
    sheet.append(
        [
            "商品名称*",
            "商品分类*",
            "商品描述",
            "SKU编码*",
            "规格_颜色*",
            "规格_容量",
            "起订量*",
            "币种*",
            "库存数量*",
            "状态*",
            "阶梯价*",
            "商品图片",
            "商品详情图",
            "SKU图片",
        ]
    )
    sheet.append(
        [
            "Sample Phone Case",
            "Accessories/Phone Cases",
            "Protective wholesale case",
            "SAMPLE-CASE-BLACK",
            "Black",
            "",
            10,
            "USD",
            100,
            "active",
            5.5,
            "https://example.com/product.jpg",
            "",
            "https://example.com/sku.jpg",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _parse_workbook(data: bytes) -> tuple[list[ProductImportRow], list[ImportError], int]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    header_index: int | None = None
    header_map: dict[int, str] = {}
    spec_columns: dict[int, str] = {}
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=10, values_only=True),
        start=1,
    ):
        values = [str(value).strip() if value is not None else "" for value in row]
        if not any("商品名称" in value or "SKU编码" in value for value in values):
            continue
        header_index = row_index
        for column_index, raw_value in enumerate(values):
            value = raw_value.rstrip("*")
            if value in FIXED_COLUMNS:
                header_map[column_index] = FIXED_COLUMNS[value]
            elif value.startswith("规格_") and value[3:]:
                spec_columns[column_index] = value[3:]
        break
    if header_index is None:
        workbook.close()
        return [], [ImportError(row=0, errors=["Template header row was not found"])], 0

    raw_rows = [
        row
        for row in sheet.iter_rows(min_row=header_index + 1, values_only=True)
        if row and any(value is not None and str(value).strip() for value in row)
    ]
    if len(raw_rows) > MAX_IMPORT_ROWS:
        workbook.close()
        return (
            [],
            [ImportError(row=0, errors=[f"Import cannot exceed {MAX_IMPORT_ROWS} rows"])],
            len(raw_rows),
        )

    valid_rows: list[ProductImportRow] = []
    errors: list[ImportError] = []
    seen_sku_codes: set[str] = set()
    for logical_row, row in enumerate(raw_rows, start=1):
        parsed: dict = {
            "row": logical_row,
            "product_name": "",
            "category_path": "",
            "description": "",
            "sku_code": "",
            "specs": {},
            "moq": 1,
            "currency": "USD",
            "stock_quantity": 0,
            "status": "active",
            "unit_price": 0,
            "product_images": [],
            "product_detail_image": "",
            "sku_image": "",
        }
        for column_index, field_name in header_map.items():
            value = row[column_index] if column_index < len(row) else None
            string_value = str(value).strip() if value is not None else ""
            if field_name in {"moq", "stock_quantity"}:
                parsed[field_name] = value if value not in {None, ""} else parsed[field_name]
            elif field_name == "unit_price":
                parsed[field_name] = value if value not in {None, ""} else 0
            elif field_name == "currency":
                parsed[field_name] = string_value.upper() or "USD"
            elif field_name == "status":
                parsed[field_name] = string_value.lower() or "active"
            elif field_name == "product_images":
                parsed[field_name] = [
                    item.strip() for item in string_value.split(",") if item.strip()
                ][:5]
            else:
                parsed[field_name] = string_value
        for column_index, spec_name in spec_columns.items():
            value = row[column_index] if column_index < len(row) else None
            string_value = str(value).strip() if value is not None else ""
            if string_value:
                parsed["specs"][spec_name] = string_value
        try:
            item = ProductImportRow.model_validate(parsed)
            if item.sku_code in seen_sku_codes:
                raise ValueError("SKU code is duplicated in the workbook")
            seen_sku_codes.add(item.sku_code)
            valid_rows.append(item)
        except (ValidationError, ValueError) as exc:
            messages = (
                [str(error["msg"]) for error in exc.errors()]
                if isinstance(exc, ValidationError)
                else [str(exc)]
            )
            errors.append(
                ImportError(
                    row=logical_row,
                    sku_code=str(parsed.get("sku_code", "")),
                    errors=messages,
                )
            )
    workbook.close()
    return valid_rows, errors, len(raw_rows)


async def _resolve_category_path(
    workspace_id: str,
    category_path: str,
    session: SessionDep,
) -> Category:
    parent_id: str | None = None
    category: Category | None = None
    for name in [part.strip() for part in category_path.split("/") if part.strip()]:
        parent_filter = (
            Category.parent_id == parent_id
            if parent_id is not None
            else Category.parent_id.is_(None)
        )
        category = await session.scalar(
            select(Category).where(
                Category.workspace_id == workspace_id,
                Category.name == name,
                parent_filter,
            )
        )
        if category is None:
            category = Category(
                id=new_id(),
                workspace_id=workspace_id,
                parent_id=parent_id,
                name=name,
            )
            session.add(category)
            await session.flush()
        parent_id = category.id
    if category is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Category path is empty",
        )
    return category


@router.get("/template")
async def download_product_import_template(context: WorkspaceDep) -> StreamingResponse:
    del context
    content = await anyio.to_thread.run_sync(_template_bytes)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="product_import_template.xlsx"'},
    )


@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_product_import(
    context: WorkspaceWriteDep,
    session: SessionDep,
    file: Annotated[UploadFile, File()],
) -> ImportPreviewResponse:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Only .xlsx files are supported",
        )
    data = await file.read(MAX_IMPORT_BYTES + 1)
    await file.close()
    if len(data) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Import file exceeds 10 MB",
        )
    try:
        rows, errors, total = await anyio.to_thread.run_sync(_parse_workbook, data)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="The workbook could not be parsed",
        ) from exc

    if rows:
        existing_codes = set(
            await session.scalars(
                select(ProductVariant.sku_code).where(
                    ProductVariant.workspace_id == context.workspace.id,
                    ProductVariant.sku_code.in_([item.sku_code for item in rows]),
                )
            )
        )
        accepted: list[ProductImportRow] = []
        for item in rows:
            if item.sku_code in existing_codes:
                errors.append(
                    ImportError(
                        row=item.row,
                        sku_code=item.sku_code,
                        errors=["SKU code already exists in this Workspace"],
                    )
                )
            else:
                accepted.append(item)
        rows = accepted
    return ImportPreviewResponse(
        rows=rows,
        errors=errors,
        summary=ImportSummary(
            total=total,
            valid=len(rows),
            invalid=len(errors),
            products=len({item.product_name for item in rows}),
            skus=len(rows),
        ),
    )


async def _import_product_batch(
    batch_rows: list[ProductImportRow],
    context: WorkspaceWriteDep,
    session: SessionDep,
) -> tuple[int, int, list[str]]:
    groups: dict[str, list[ProductImportRow]] = defaultdict(list)
    created_product_ids: list[str] = []
    for row in batch_rows:
        groups[row.product_name].append(row)
    for product_name, product_rows in groups.items():
        first = product_rows[0]
        category = await _resolve_category_path(
            context.workspace.id,
            first.category_path,
            session,
        )
        specification_values: dict[str, set[str]] = defaultdict(set)
        for row in product_rows:
            for key, value in row.specs.items():
                if value:
                    specification_values[key].add(value)
        product = Product(
            id=new_id(),
            workspace_id=context.workspace.id,
            category_id=category.id,
            name=product_name,
            description=first.description,
            specification_template=[
                {"name": key, "options": sorted(values)}
                for key, values in specification_values.items()
            ],
            status=first.status,
        )
        created_product_ids.append(product.id)
        session.add(product)
        for url in first.product_images:
            session.add(
                ProductImage(
                    id=new_id(),
                    product_id=product.id,
                    image_type="product",
                    file_key=url,
                )
            )
        if first.product_detail_image:
            session.add(
                ProductImage(
                    id=new_id(),
                    product_id=product.id,
                    image_type="product_detail",
                    file_key=first.product_detail_image,
                )
            )
        for row in product_rows:
            variant = ProductVariant(
                id=new_id(),
                workspace_id=context.workspace.id,
                product_id=product.id,
                sku_code=row.sku_code,
                specifications=row.specs,
                minimum_order_quantity=row.moq,
                currency=row.currency,
                stock_quantity=row.stock_quantity,
                status=row.status,
            )
            session.add(variant)
            session.add(
                PriceTier(
                    id=new_id(),
                    variant_id=variant.id,
                    minimum_quantity=row.moq,
                    unit_price=row.unit_price,
                )
            )
            if row.sku_image:
                session.add(
                    ProductImage(
                        id=new_id(),
                        variant_id=variant.id,
                        image_type="sku",
                        file_key=row.sku_image,
                    )
                )
    record_audit(
        session,
        workspace_id=context.workspace.id,
        actor_user_id=context.user.id,
        action="catalog.products_imported",
        entity_type="workspace",
        entity_id=context.workspace.id,
        payload={"product_count": len(groups), "variant_count": len(batch_rows)},
    )
    await enqueue_shopify_sync_if_active(
        session,
        workspace_id=context.workspace.id,
        operation="catalog",
        product_ids=created_product_ids,
    )
    return len(groups), len(batch_rows), created_product_ids


@router.post("/confirm", response_model=ImportConfirmResponse, status_code=status.HTTP_201_CREATED)
async def confirm_product_import(
    payload: ImportConfirmRequest,
    context: WorkspaceWriteDep,
    session: SessionDep,
) -> ImportConfirmResponse:
    sku_codes = [item.sku_code for item in payload.rows]
    if len(sku_codes) != len(set(sku_codes)):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Import rows contain duplicate SKU codes",
        )
    existing = set(
        await session.scalars(
            select(ProductVariant.sku_code).where(
                ProductVariant.workspace_id == context.workspace.id,
                ProductVariant.sku_code.in_(sku_codes),
            )
        )
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"message": "SKU codes already exist", "sku_codes": sorted(existing)},
        )

    created_products = 0
    created_skus = 0
    batches = 0
    errors: list[str] = []
    for offset in range(0, len(payload.rows), IMPORT_BATCH_SIZE):
        batches += 1
        batch = payload.rows[offset : offset + IMPORT_BATCH_SIZE]
        try:
            product_count, sku_count, _product_ids = await _import_product_batch(
                batch,
                context,
                session,
            )
            await session.commit()
            created_products += product_count
            created_skus += sku_count
        except Exception as exc:
            await session.rollback()
            error_kind = "catalog conflict" if isinstance(exc, IntegrityError) else "write failure"
            errors.append(f"Batch {batches} failed due to {error_kind}")
    return ImportConfirmResponse(
        created_products=created_products,
        created_skus=created_skus,
        batches=batches,
        errors=errors,
    )
